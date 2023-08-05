import os
import openpyxl
import re
from collections import defaultdict
from openpyxl import Workbook

def get_excel_files_in_directory(directory_path):
    excel_files = []
    for root, _, files in os.walk(directory_path):
        for file in files:
            if file.endswith('.xlsx'):
                excel_files.append(os.path.join(root, file))

    if excel_files:
        print("カレントディレクトリ以下のエクセルファイルのパス：")
        for file_path in excel_files:
            print(file_path)
    else:
        print("カレントディレクトリ以下にエクセルファイルが見つかりませんでした。")
    return excel_files

def read_rows_with_decision_status(file_paths):
    aggregation = []
    for file_path in file_paths:
        try:
            wb =openpyxl.load_workbook(file_path)
            sheet_names = wb.sheetnames
            alphabet_sheet_names = [name for name in sheet_names if re.match(r'^[a-zA-Z0-9]+$', name)]
            #print(alphabet_sheet_names)
            for sheet_name in alphabet_sheet_names:
                sheet = wb[sheet_name]
                reference_point = serch_reference_point(sheet)
                aggregation = aggregation + get_test_data(sheet,reference_point)
                
        except Exception as e:
            print(f"ファイル '{file_path}' の読み込み中にエラーが発生しました：{e}")
    print(aggregation_test_data(aggregation))
    OK_per_date = aggregate_OK(aggregation)
    print(aggregate_by_date(OK_per_date))
    write_data_to_sheet(os.path.join(os.getcwd(), "path_to_output_excel_file.xlsx"),"test",aggregation_test_data(aggregation))
    write_datalist_to_sheet(os.path.join(os.getcwd(), "path_to_output_excel_file.xlsx"),"AAAAA",aggregate_by_date(OK_per_date))



def write_data_to_sheet(file_path, sheet_name, data):
    print(file_path)
    # 新しいワークブックを作成
    wb =openpyxl.load_workbook(file_path)

    # ワークブックから指定したシートを取得
    ws = wb[sheet_name]

    # キーをエクセルの列に、バリューを行に対応させて書き込む
    for row_idx, (key, value) in enumerate(data.items(), start=1):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)

    # ワークブックを保存
    wb.save(file_path)

def write_datalist_to_sheet(file_path, sheet_name, data_list):
    # 新しいワークブックを作成
    wb =openpyxl.load_workbook(file_path)

    # ワークブックから指定したシートを取得
    ws = wb[sheet_name]

    # ヘッダ行を書き込む
    header_row = 1
    for col_idx, key in enumerate(data_list[0].keys(), start=1):
        ws.cell(row=header_row, column=col_idx, value=key)

    # データを書き込む
    for row_idx, data in enumerate(data_list, start=2):
        for col_idx, value in enumerate(data.values(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(file_path)




def aggregate_OK(aggregation):
    list = []
    for data in aggregation:
        if(data["judgement"]is not None):
            flag = 0
            if(data["judgement"]!="不合格"):
                flag = 1
            test = {
                "judgement_num":flag,
                "date":data["date"]
            }
        list.append(test)
    #print(list)
    return list

def aggregate_by_date(data_list):
    # 日付ごとの集計結果を格納する辞書
    date_aggregation = defaultdict(int)
    for data in data_list:
        date = data["date"]
        value = data["judgement_num"]
        date_aggregation[date] += value
    result = [{"date": date, "judgement_num": date_aggregation[date]} for date in date_aggregation]
    return result


def serch_reference_point(sheet):
    #print("START_serch_reference_point")
    for row in range(1,100):
        if (sheet.cell(row=row, column=1).value=="試験対象"):
            #print(sheet.cell(row=row, column=2).value)
            return row
        
def get_test_data(sheet,reference_point):
    #print("START_get_test_data")
    test_list_method =[]  
    for col in range(1,100):
        if (sheet.cell(row=reference_point, column=col).value=="◯"):
            test = {
                "judgement":sheet.cell(row=reference_point+1, column=col).value,
                "date": sheet.cell(row=reference_point+2, column=col).value
            }
            test_list_method.append(test)
    #print(test_list_method)
    return test_list_method

def aggregation_test_data(test_list_method):
    num_tests_performed = 0
    num_OK = 0
    for data in test_list_method:
        if(data["judgement"]is not None):
            num_tests_performed += 1
            if(data["judgement"]!="不合格"):
                num_OK += 1
    aggregation_data = {
        "num_test_items" :len(test_list_method),
        "num_tests_performed" : num_tests_performed,
        "num_OK" : num_OK
    }
    return aggregation_data
        

if __name__ == "__main__":
    # カレントディレクトリを取得
    current_directory = os.getcwd()

    # カレントディレクトリ以下のエクセルファイルのパスを取得
    excel_files_list = get_excel_files_in_directory(current_directory)

    # 条件に合う値を読み込む
    decision_rows_data = read_rows_with_decision_status(excel_files_list)



